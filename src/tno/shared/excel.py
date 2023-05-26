import base64
from collections import OrderedDict
import datetime
from decimal import Decimal
from enum import Enum
import io
from typing import Dict, Iterable

from openpyxl import Workbook
from openpyxl.styles import Font, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


default_styles_as_dict: Dict[str, NamedStyle] = {
    "decimal": NamedStyle("decimal", number_format="#,##0.00"),
    "date": NamedStyle("date", number_format="dd/mm/yyyy"),
    "column_header": NamedStyle("column_header", font=Font(bold=True)),
    "normal": NamedStyle("normal"),
}


def default_styles() -> Iterable[NamedStyle]:
    """
    Add several NamedStyles to the workbook. This way a cell can be styled with the name:

    sheet["A3"].style = "currency_EUR"

    Add these styles to a Workbook instance by invoking workbook.add_named_style().
    """
    return default_styles_as_dict.values()


def base64encode_excel_file(wb: Workbook) -> str:
    """
    Encodes an excel file as base64 to send it back to the frontend.
    """
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return base64.b64encode(stream.read()).decode()


def create_workbook(sheet_name: str) -> Workbook:
    """
    Create a workbook with the first sheet named sheet_name and the styles registered.

    Args:
         - sheet_name: Name of the first sheet
         - named_styles: A list of NamedStyle that can be used for formatting cells.

    Returns:
        A workbook instance.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for named_style in default_styles():
        workbook.add_named_style(named_style)
    return workbook


def create_simple_excel_file(
    sheet_name: str, field_map: OrderedDict, entities: Iterable
) -> Workbook:
    """
    Create a simple excel file, with all entities on a single sheet.

    The field_map should be an OrderedDict, with as keys the column headers, and
    as values the dot-separated attributes. Example:

    field_map = collections.OrderedDict()
    field_map["BAG ID"] = "bag_id"
    field_map["Bedrijfsnaam"] = "company.name"
    """
    workbook = create_workbook(sheet_name=sheet_name)
    sheet = workbook.active
    write_sheet_header(sheet, field_map)

    for row, entity in enumerate(entities, start=2):
        write_sheet_row(sheet, entity, row, field_map)
    return workbook


def add_excel_sheet(
    workbook: Workbook, sheet_name: str, field_map: OrderedDict, entities: Iterable
) -> Workbook:
    sheet = workbook.create_sheet(sheet_name)

    write_sheet_header(sheet, field_map)

    for row, entity in enumerate(entities, start=2):
        write_sheet_row(sheet, entity, row, field_map)
    return workbook


def write_sheet_header(sheet: Worksheet, field_map: OrderedDict) -> None:
    """
    Write the header for the sheet on the specified sheet.
    The header looks like:
    Rank | Total |
    """
    col = 1
    for header_col in field_map.keys():
        cell = sheet[f"{get_column_letter(col)}1"]
        cell.value = header_col
        cell.style = "column_header"

        # If it is specified in the field_map: set the width of the column.
        field = field_map[header_col]
        if type(field) is dict and "width" in field:
            sheet.column_dimensions[get_column_letter(col)].width = field["width"]
        else:
            sheet.column_dimensions[get_column_letter(col)].width = 12
        col += 1


def write_empty_sheet_row(sheet: Worksheet, row: int, number_of_cols: int) -> None:
    """
    Write an empty line in the sheet.
    """
    for col in range(number_of_cols):
        sheet[f"{get_column_letter(col + 1)}{row}"].value = ""


def write_sheet_row(
    sheet: Worksheet,
    entity,
    row: int,
    field_map: OrderedDict,
) -> None:
    """
    Write a line in the sheet.

    Args:
        sheet: Worksheet instance on which the row is written.
        entity (Object): An object or a dict containing the entity for which
            the row is written.
        row: Row number to write to.
        field_map: An ordered dict of which the values determine
            what information to write and can be a dict, a callable or a key in
            the entity.
    """

    col = 1
    for field in field_map.values():
        value = getattrd(entity, field)

        # For enum values get the enum.value.
        if isinstance(value, Enum):
            value = value.value

        cell = sheet[f"{get_column_letter(col)}{row}"]

        # Format dates.
        if type(value) in (datetime.datetime, datetime.date):
            if type(value) == datetime.datetime:
                value = value.replace(tzinfo=None)
            cell.value = value
            cell.style = "date"
        elif type(value) == Decimal:
            value = float(value)
            cell.value = value
            cell.style = "decimal"
        else:
            cell.value = value
            cell.style = "normal"

        col += 1


def getattrd(obj, name: str):
    """
    Same as getattr(), but allows dot notation lookup and works on dicts.
    """
    # Starting point == obj
    value = obj
    if isinstance(value, dict):
        for n in name.split("."):
            value = value.get(n)
    else:
        for n in name.split("."):
            value = getattr(value, n)
    return value
