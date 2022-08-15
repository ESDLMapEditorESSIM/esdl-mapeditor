import json
import sys

import esdl
from esdl.esdl_handler import EnergySystemHandler
from influxdb import InfluxDBClient
from influxdb.resultset import ResultSet
from openpyxl import Workbook

from extensions.dice_workflow.util import all_energy_assets, all_buildings, all_energy_assets_from_area


class EnergyHolder:
    def __init__(self):
        self.positive = 0
        self.negative = 0

    def add(self, val):
        if val >= 0:
            self.positive += val
        else:
            self.negative += val


class SummedAsset:
    def __init__(self, asset_name, asset_class, capability, group):
        self._id = f"{asset_name}-{asset_class}-{group}"
        self.asset_name = asset_name
        self.asset_class = asset_class
        self.capability = capability
        self.carriers = {}
        self.power = 0
        self.capacity = 0
        self.group = group

    def id(self):
        return self._id

    def plus_carrier(self, carrier_name: str, allocation_energy: float):
        if carrier_name in self.carriers:
            self.carriers[carrier_name].add(allocation_energy)
        else:
            e = EnergyHolder()
            e.add(allocation_energy)
            self.carriers[carrier_name] = e

    def plus_power(self, power):
        self.power += power

    def plus_capacity(self, capacity):
        self.capacity += capacity

# class LazyDict():
#     def __init__(self, merge_func, init_func=None, d=None):
#         if d is not None:
#             self.d = d
#         else:
#             self.d = {}
#         self.init_func = init_func
#         self.merge_func = merge_func
#
#     def add(self, key, value):
#         if key in self.d:
#             self.d[key] = self.merge_func(self.d[key], value)
#         else:
#             self.d[key] = value


JOULE_TO_KWH = "0.000000277777777777778"
GROOTVERBRUIKER_KWH = 50000 #KWh used


def get_big_consumers(es: esdl.EnergySystem):
    buildings = all_buildings(es.instance[0])
    groot = []
    for building in buildings:
        building: esdl.Building
        for asset in building.asset:
            if isinstance(asset, esdl.EConnection):
                if asset.assetType == "GVB":
                    groot.append(building.id)
    return groot

    # query = f"SELECT sum(\"allocationEnergy\") * {JOULE_TO_KWH} FROM  /.*/ WHERE (\"simulationRun\" = '{simulation_run}' AND \"assetClass\" = 'ElectricityDemand') AND time >= 1546297200000ms and time <= 1577836800000ms GROUP BY \"address\""
    # grootverbruikers = []
    # kleinverbuikers = []
    # result: ResultSet = influx_client.query(query, database=es.id)
    # for info, values in result.items():
    #     prop = info[1]
    #     address = prop["address"]
    #     for val in values:
    #         if val["sum"] >= GROOTVERBRUIKER_KWH:
    #             grootverbruikers.append(address)
    #         else:
    #             kleinverbuikers.append(address)
    # return grootverbruikers, kleinverbuikers


def regex_list(l: list):
    s = "/"
    first = True
    for e in l:
        if not first:
            s += "|"
        s += e
    return s + "/"


def get_group(asset: esdl.EnergyAsset, big_consumers) -> str:
    if asset.containingBuilding is not None:
        for port in asset.port:
            if port.carrier.name == 'Elektriciteit':
                if asset.containingBuilding.id in big_consumers:
                    return "GVB"
                else:
                    return "KVB"
    else:
        return "field"

def year_total_per_asset_class(influx_client: InfluxDBClient, simulation_run: str, es: esdl.EnergySystem, big_consumers: list, hashed_energy_assets):
    summed_assets = {}
    saldering_building = {}
    query = f"SELECT sum(\"allocationEnergy\") * {JOULE_TO_KWH} FROM /.*/ WHERE (\"simulationRun\" = '{simulation_run}') AND time >= 1546297200000ms and time <= 1577836800000ms and \"allocationEnergy\" > 0 GROUP BY \"assetId\", \"assetName\", \"assetClass\", \"carrierName\", \"capability\"; " \
            f"SELECT sum(\"allocationEnergy\") * {JOULE_TO_KWH} FROM /.*/ WHERE (\"simulationRun\" = '{simulation_run}') AND time >= 1546297200000ms and time <= 1577836800000ms and \"allocationEnergy\" < 0 GROUP BY \"assetId\", \"assetName\", \"assetClass\", \"carrierName\", \"capability\""
    result: ResultSet = influx_client.query(query, database=es.id)
    for sign in result:
        for info, values in sign.items():
            prop = info[1]
            asset: esdl.EnergyAsset = hashed_energy_assets[prop['assetId']]
            group = get_group(asset, big_consumers)

            summed_asset = SummedAsset(asset.name, prop['assetClass'], prop['capability'], group)

            for val in values:
                allocation_energy = val['sum']
                summed_asset_id = summed_asset.id()
                if summed_asset_id not in summed_assets:
                    summed_assets[summed_asset_id] = summed_asset
                summed_assets[summed_asset_id].plus_carrier(prop['carrierName'], allocation_energy)

                # added for salderings calculation
                if isinstance(asset, esdl.EConnection):
                    if group == "KVB":
                        building_id = asset.containingBuilding.id
                        if building_id not in saldering_building:
                            saldering_building[building_id] = EnergyHolder()
                        saldering_building[building_id].add(allocation_energy)
    # sum saldering
    compensated = 0.0
    to_pay = 0.0
    overproduction = 0.0
    for holder in saldering_building.values():
        provided_back = holder.negative*-1
        imported = holder.positive

        compensated += min(provided_back, imported)
        to_pay += max(imported-provided_back, 0)
        overproduction += max(provided_back-imported, 0)
    return summed_assets, {"compensated": compensated, "overproduction": overproduction, "to_pay": to_pay}


def get_max_profile_influx(influx_client: InfluxDBClient, profile: esdl.InfluxDBProfile):
    query = f"SELECT max(\"{profile.field}\") FROM \"{profile.measurement}\" WHERE time >= 1546297200000ms and time <= 1577833201000ms"
    result: ResultSet = influx_client.query(query, database=profile.database)
    for info, values in result.items():
        for val in values:
            return profile.multiplier * val["max"]


def get_power(influx_client: InfluxDBClient, asset):
    # if the power is set in the asset, then use that value.
    if not asset.power == 0.0:
        return asset.power
    return 0.0
    # for port in producer.port:
    #     port: esdl.Port
    #     for profile in port.profile:
    #         profile: esdl.GenericProfile
    #         if isinstance(profile, esdl.InfluxDBProfile):
    #             profile: esdl.InfluxDBProfile
    #             return get_max_profile_influx(influx_client, profile)


def extract_power(instance: esdl.Instance, summed_assets, big_consumers):
    energy_assets = all_energy_assets(instance)
    for energy_asset in energy_assets:
        energy_asset: esdl.EnergyAsset
        group = get_group(energy_asset, big_consumers)
        ea_summed_id = SummedAsset(energy_asset.name, type(energy_asset).__name__, "", group).id()
        if ea_summed_id in summed_assets:
            summed_asset = summed_assets[ea_summed_id]
            if isinstance(energy_asset, esdl.Consumer):
                energy_asset: esdl.Consumer
                summed_asset.plus_power(energy_asset.power)
            elif isinstance(energy_asset, esdl.Producer):
                energy_asset: esdl.Producer
                summed_asset.plus_power(energy_asset.power)
            elif isinstance(energy_asset, esdl.AbstractBasicConversion):
                energy_asset: esdl.AbstractBasicConversion
                summed_asset.plus_power(energy_asset.power)


def extract_capacity(instance: esdl.Instance, summed_assets, big_consumers):
    energy_assets = all_energy_assets(instance)
    for energy_asset in energy_assets:
        # energy_asset: esdl.EnergyAsset
        if isinstance(energy_asset, esdl.Storage):
            energy_asset: esdl.Storage
            group = get_group(energy_asset, big_consumers)
            ea_summed_id = SummedAsset(energy_asset.name, type(energy_asset).__name__, "", group).id()
            summed_asset = summed_assets[ea_summed_id]
            summed_asset.plus_capacity(energy_asset.capacity * float(JOULE_TO_KWH))


def extract_energy_asset_classes(bc):
    energy_asset_classes = {}
    for summed_assets in bc.values():
        for summed_asset in summed_assets.values():
            if summed_asset.id() not in energy_asset_classes:
                energy_asset_classes[summed_asset.id()] = summed_asset
    return list(energy_asset_classes.values())


# Checks if the cell if empty before it is filled
def safe_fill_cell(sheet, row, column, val):
    if not sheet.cell(row, column).value is None:
        raise Exception(f"the cell that was supposed to be filled is already full, row {row} column {column}")
    sheet.cell(row, column, val)


def bc_export_excel(bc, salderings) -> Workbook:
    energy_asset_classes = extract_energy_asset_classes(bc)

    excel_dict = {"c": {}, "r": {}}
    book = Workbook()
    sheet = book.active

    excel_dict["r"]["esdl_name"] = 1
    excel_dict["r"]["column_name"] = 2
    excel_dict["c"]["energy_asset_name"] = 1
    safe_fill_cell(sheet, excel_dict["r"]["column_name"], excel_dict["c"]["energy_asset_name"], "asset_name")
    excel_dict["c"]["energy_asset_type"] = 2
    safe_fill_cell(sheet, excel_dict["r"]["column_name"], excel_dict["c"]["energy_asset_type"], "asset_class")
    excel_dict["c"]["energy_asset_capability"] = 3
    safe_fill_cell(sheet, excel_dict["r"]["column_name"], excel_dict["c"]["energy_asset_capability"], "capability")
    excel_dict["c"]["energy_asset_group"] = 4
    safe_fill_cell(sheet, excel_dict["r"]["column_name"], excel_dict["c"]["energy_asset_group"], "group")
    excel_dict["c"]["carrier"] = 5
    safe_fill_cell(sheet, excel_dict["r"]["column_name"], excel_dict["c"]["carrier"], "carrier")

    row_counter = len(excel_dict["r"])+1

    # asset names
    for summed_asset in energy_asset_classes:
        excel_dict["r"][summed_asset.id()] = row_counter

        column = excel_dict["c"]["energy_asset_name"]
        safe_fill_cell(sheet, row_counter, column, summed_asset.asset_name)
        column = excel_dict["c"]["energy_asset_type"]
        safe_fill_cell(sheet, row_counter, column, summed_asset.asset_class)
        column = excel_dict["c"]["energy_asset_group"]
        safe_fill_cell(sheet, row_counter, column, summed_asset.group)
        column = excel_dict["c"]["energy_asset_capability"]
        safe_fill_cell(sheet, row_counter, column, summed_asset.capability)
        first = True
        for carrier, energy_holder in summed_asset.carriers.items():
            if not first:
                row_counter += 1
            column = excel_dict["c"]["carrier"]
            safe_fill_cell(sheet, row_counter, column, carrier)
            # need to check if maybe in some other business cases there are negative and positive values
            for bc_summed_assets in bc.values():
                if summed_asset.id() in bc_summed_assets:
                    bc_energy_holder = bc_summed_assets[summed_asset.id()].carriers[carrier]
                    # need an extra line to place positive and negative value
                    if bc_energy_holder.positive > 0 and bc_energy_holder.negative < 0:
                        row_counter += 1
                        safe_fill_cell(sheet, row_counter, column, carrier)
                        break
            first = False
        row_counter += 1

    # Add saldering lines with empty line above to distantiate it
    row_counter += 1
    safe_fill_cell(sheet, row_counter, 1, "Gesaldeerde teruggeleverde Elektriciteit")
    excel_dict["r"]["compensated"] = row_counter
    row_counter += 1
    safe_fill_cell(sheet, row_counter, 1, "Ongesaldeerde teruggeleverde Elektriciteit")
    excel_dict["r"]["overproduction"] = row_counter
    row_counter += 1
    safe_fill_cell(sheet, row_counter, 1, "Te betalen geconsumeerde Elektriciteit")
    excel_dict["r"]["to_pay"] = row_counter
    row_counter += 1

    column_counter = len(excel_dict["c"])+1
    for esdl_name, summed_assets in bc.items():
        safe_fill_cell(sheet, excel_dict["r"]["esdl_name"], column_counter, esdl_name)

        # total_year
        safe_fill_cell(sheet, excel_dict["r"]["column_name"], column_counter, "Energy (kWh/year)")
        for summed_asset_id, summed_asset in summed_assets.items():
            for sub_row, energy_holder in enumerate(summed_asset.carriers.values()):
                sub_sub_row = 0
                if energy_holder.positive > 0:
                    safe_fill_cell(sheet, excel_dict["r"][summed_asset_id] + sub_row, column_counter, energy_holder.positive)
                    sub_sub_row += 1
                if energy_holder.negative < 0:
                    safe_fill_cell(sheet, excel_dict["r"][summed_asset_id] + sub_row + sub_sub_row, column_counter, energy_holder.negative)

        # saldering
        saldering = salderings[esdl_name]
        safe_fill_cell(sheet, excel_dict["r"]["compensated"], column_counter, saldering["compensated"])
        safe_fill_cell(sheet, excel_dict["r"]["overproduction"], column_counter, saldering["overproduction"])
        safe_fill_cell(sheet, excel_dict["r"]["to_pay"], column_counter, saldering["to_pay"])

        column_counter += 1

        # power
        safe_fill_cell(sheet, excel_dict["r"]["column_name"], column_counter, "Power (W)")
        for summed_asset_id, summed_asset in summed_assets.items():
            # for sub_row, year_total in enumerate(summed_asset.carriers.values()):
            if summed_asset.power > 0:
                safe_fill_cell(sheet, excel_dict["r"][summed_asset_id], column_counter, summed_asset.power)
        column_counter += 1

        # capacity
        safe_fill_cell(sheet, excel_dict["r"]["column_name"], column_counter, "Capacity (kWh)")
        for summed_asset_id, summed_asset in summed_assets.items():
            # for sub_row, year_total in enumerate(summed_asset.carriers.values()):
            if summed_asset.capacity > 0:
                safe_fill_cell(sheet, excel_dict["r"][summed_asset_id], column_counter, summed_asset.capacity)
        column_counter += 1



    #Saldering


    return book


def _bc_export_excel(bc, salderings, output_excel_name: str):
    book = bc_export_excel(bc, salderings)
    book.save(output_excel_name)


def sum_assets_single_case(influx_client: InfluxDBClient, simulation_run: str, es: esdl.EnergySystem):
    print(f"Getting data for: {es.name}")
    influx_client.switch_database(es.id)
    big_consumers = get_big_consumers(es)
    hashed_energy_assets = {x.id: x for x in all_energy_assets_from_area(es.instance[0].area) if isinstance(x, esdl.EnergyAsset)}
    summed_assets, saldering = year_total_per_asset_class(influx_client, simulation_run, es, big_consumers, hashed_energy_assets)
    # summed_assets = {}
    extract_power(es.instance[0], summed_assets, big_consumers)
    extract_capacity(es.instance[0], summed_assets, big_consumers)
    return summed_assets, saldering


def main():
    if not len(sys.argv) == 2:
        print("usage: python3 export_bc.py <path to export.json>")
        exit(-1)
    with open(sys.argv[1]) as f:
        export_json = json.load(f)

    esh = EnergySystemHandler()

    influx_client = InfluxDBClient(export_json["influxdb"]["url"], export_json["influxdb"]["port"])

    bc = {}
    salderings = {}
    for export in export_json["exports"]:
        es, _ = esh.load_file(export["esdl_file_case"])
        simulation_run = export["simulation_run_case"]
        bc[es.name], salderings[es.name] = sum_assets_single_case(influx_client, simulation_run, es)
    _bc_export_excel(bc, salderings, export_json["output_excel"])
    influx_client.close()


def export_business_case(
    influx_client: InfluxDBClient,
    essim_id: str,
    es: esdl.EnergySystem,
) -> Workbook:
    """Entrypoint for the mapeditor."""
    bc = dict()
    bc[es.name] = sum_assets_single_case(influx_client, essim_id, es)
    return bc_export_excel(bc)


if __name__ == '__main__':
    main()
